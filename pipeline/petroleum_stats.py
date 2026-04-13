"""Parse Australian Petroleum Statistics Excel files from data.gov.au."""

import os
import requests
from openpyxl import load_workbook

DATASET_URL = "https://data.gov.au/data/api/3/action/package_show?id=australian-petroleum-statistics"


def download_latest_excel(output_path: str) -> str:
    resp = requests.get(DATASET_URL, timeout=30)
    resp.raise_for_status()
    resources = resp.json()["result"]["resources"]
    xlsx_resources = [
        r for r in resources
        if r["format"].upper() in ("XLSX", "XLS") or r["url"].endswith(".xlsx")
    ]
    if not xlsx_resources:
        raise RuntimeError("No Excel resource found in dataset")
    resource = xlsx_resources[0]
    download_url = resource["url"]
    data_resp = requests.get(download_url, timeout=120)
    data_resp.raise_for_status()
    with open(output_path, "wb") as f:
        f.write(data_resp.content)
    return output_path


def parse_imports_sheet(excel_path: str) -> list[dict]:
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb["Imports volume"]
    records = []
    header_row = None

    for row in ws.iter_rows(values_only=True):
        if row[0] == "Month":
            header_row = row
            continue
        if header_row is None:
            continue
        month_val = row[0]
        if month_val is None:
            continue
        month_str = month_val.strftime("%Y-%m") if hasattr(month_val, "strftime") else str(month_val)

        def safe_float(val):
            if val is None or val == "n.a." or val == "":
                return 0.0
            try:
                return float(val)
            except (ValueError, TypeError):
                return 0.0

        # Column layout from the actual Excel:
        # 0: Month
        # 1: Crude oil & other refinery feedstocks (ML)
        # 2: LPG (ML)
        # 3: Automotive gasoline (ML)
        # 4: Aviation gasoline (ML)
        # 5: Aviation turbine fuel (ML)
        # 6: Kerosene & heating oil (ML)
        # 7: Diesel oil (ML)
        # 8: Fuel oil (ML)
        # 9: Lubricating oils, greases & basestocks (ML)
        # 10: Bitumen (ML)
        # 11: Other products (ML)
        # 12: Total refined petroleum products (ML)
        # 13: Total oil imports (ML)
        # 14: Natural gas (ML)
        record = {
            "month": month_str,
            "crude_oil_ml": safe_float(row[1]),
            "lpg_ml": safe_float(row[2]),
            "gasoline_ml": safe_float(row[3]),
            "jet_fuel_ml": safe_float(row[5]),
            "diesel_ml": safe_float(row[7]),
            "fuel_oil_ml": safe_float(row[8]),
            "total_ml": safe_float(row[13]),
        }
        records.append(record)

    wb.close()
    return records


def parse_consumption_cover(excel_path: str) -> list[dict]:
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb["Consumption cover"]
    records = []
    header_row = None

    for row in ws.iter_rows(values_only=True):
        if row[0] == "Month":
            header_row = row
            continue
        if header_row is None:
            continue
        month_val = row[0]
        if month_val is None:
            continue
        month_str = month_val.strftime("%Y-%m") if hasattr(month_val, "strftime") else str(month_val)

        def safe_int(val):
            if val is None or val == "n.a." or val == "":
                return 0
            try:
                return int(val)
            except (ValueError, TypeError):
                return 0

        # Column layout:
        # 0: Month
        # 1: Crude oil and refinery feedstocks (days)
        # 2: LPG (days)
        # 3: Automotive gasoline (days)
        # 4: Aviation gasoline (days)
        # 5: Aviation turbine fuel (days)
        # 6: Diesel oil (days)
        # 7: Fuel oil (days)
        # 8: Lubricating oils, greases & basestocks (days)
        # 9: Other products (days)
        # 10: Total product stocks COE (days)
        record = {
            "month": month_str,
            "crude_days": safe_int(row[1]),
            "gasoline_days": safe_int(row[3]),
            "jet_fuel_days": safe_int(row[5]),
            "diesel_days": safe_int(row[6]),
            "total_days": safe_int(row[10]),
        }
        records.append(record)

    wb.close()
    return records


def build_imports_json(excel_path: str) -> dict:
    imports = parse_imports_sheet(excel_path)
    consumption = parse_consumption_cover(excel_path)
    return {
        "imports_by_month": imports,
        "consumption_cover": consumption,
    }
